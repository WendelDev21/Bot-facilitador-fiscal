import sys
import openpyxl
import pyperclip
import pyautogui
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QFileDialog, QMessageBox, QComboBox, QLabel, QHBoxLayout

class App(QWidget):
    def __init__(self):
        super().__init__()
        
        self.initUI()

    def initUI(self):
        self.setWindowTitle('Importar Planilha Excel')

        layout = QVBoxLayout()

        self.btnSelecionar = QPushButton('Selecionar Planilha', self)
        self.btnSelecionar.clicked.connect(self.abrir_dialogo)
        layout.addWidget(self.btnSelecionar)

        self.comboPlanilhas = QComboBox(self)
        layout.addWidget(self.comboPlanilhas)

        self.btnConfirmar = QPushButton('Confirmar Seleção', self)
        self.btnConfirmar.setEnabled(False)
        self.btnConfirmar.clicked.connect(self.confirmar_selecao)
        layout.addWidget(self.btnConfirmar)

        self.setLayout(layout)
        self.show()

    def abrir_dialogo(self):
        options = QFileDialog.Options()
        caminho_arquivo, _ = QFileDialog.getOpenFileName(self, 
            "Selecionar Planilha Excel", "", 
            "Arquivos Excel (*.xlsx);;Todos os Arquivos (*)", options=options)
        
        if caminho_arquivo:
            self.carregar_planilhas(caminho_arquivo)
        else:
            QMessageBox.warning(self, 'Aviso', 'Nenhum arquivo selecionado.')

    def carregar_planilhas(self, caminho_arquivo):
        self.caminho_arquivo = caminho_arquivo
        self.comboPlanilhas.clear()

        try:
            workbook = openpyxl.load_workbook(caminho_arquivo)
            planilhas = workbook.sheetnames
            
            if planilhas:
                self.comboPlanilhas.addItems(planilhas)
                self.btnConfirmar.setEnabled(True)
            else:
                QMessageBox.warning(self, 'Aviso', 'Nenhuma planilha encontrada no arquivo.')
                self.btnConfirmar.setEnabled(False)

        except Exception as e:
            QMessageBox.critical(self, 'Erro', f'Erro ao carregar planilhas: {e}')

    def confirmar_selecao(self):
        planilha_selecionada = self.comboPlanilhas.currentText()
        
        if planilha_selecionada:
            self.preencher_formulario(self.caminho_arquivo, planilha_selecionada)
        else:
            QMessageBox.warning(self, 'Aviso', 'Nenhuma planilha selecionada.')

    def preencher_formulario(self, caminho_arquivo, nome_planilha):
        try:
            # Entrar na planilha
            workbook = openpyxl.load_workbook(caminho_arquivo)
            sheet_teste = workbook[nome_planilha]

            # Copiar e colar as informações
            for linha in sheet_teste.iter_rows(min_row=2):
                cliente = linha[0].value
                pyautogui.click(740,229,duration=1)
                pyperclip.copy(cliente)
                pyautogui.hotkey('ctrl', 'v')

                cpf_cnpj = linha[1].value
                pyautogui.click(756,276,duration=1) 
                pyperclip.copy(cpf_cnpj) 
                pyautogui.hotkey('ctrl', 'v')

                cep = linha[2].value
                pyautogui.click(763,321,duration=1)
                pyperclip.copy(cep)
                pyautogui.hotkey('ctrl', 'v')

                endereco = linha[3].value
                pyautogui.click(765,367,duration=1)
                pyperclip.copy(endereco)
                pyautogui.hotkey('ctrl', 'v')

                bairro = linha[4].value
                pyautogui.click(769,415,duration=1)
                pyperclip.copy(bairro)
                pyautogui.hotkey('ctrl', 'v')

                municipio = linha[5].value
                pyautogui.click(770,461,duration=1)
                pyperclip.copy(municipio)
                pyautogui.hotkey('ctrl', 'v')

                uf = linha[6].value
                pyautogui.click(772,509,duration=1)
                pyperclip.copy(uf)
                pyautogui.hotkey('ctrl', 'v')

                inscricao_estadual = linha[7].value
                pyautogui.click(790,557,duration=1)
                pyperclip.copy(inscricao_estadual)
                pyautogui.hotkey('ctrl', 'v')

                descricao = linha[8].value
                pyautogui.click(792,604,duration=1)
                pyperclip.copy(descricao)
                pyautogui.hotkey('ctrl', 'v')

                quantidade = linha[9].value
                pyautogui.click(796,652,duration=1)
                pyperclip.copy(quantidade)
                pyautogui.hotkey('ctrl', 'v')

                valor_unitario = linha[10].value
                pyautogui.click(798,701,duration=1)
                pyperclip.copy(valor_unitario)
                pyautogui.hotkey('ctrl', 'v')

                valor_total = linha[11].value
                pyautogui.click(800,746,duration=1)
                pyperclip.copy(valor_total)
                pyautogui.hotkey('ctrl', 'v')

                # Clicar no botão gerar NF
                pyautogui.click(801,799,duration=1)

            QMessageBox.information(self, 'Sucesso', 'Formulário preenchido com sucesso!')

        except Exception as e:
            QMessageBox.critical(self, 'Erro', f'Erro ao preencher o formulário: {e}')

if __name__ == "__main__":
    app = QApplication(sys.argv)
    ex = App()
    sys.exit(app.exec_())
